from flask import Blueprint, render_template, redirect, url_for, flash, request, current_app, session
from flask_login import login_required, current_user
from extensions import db
from models import Upload, Result, Attendance
from processing import process, get_choice_options
from translations import get_translations
import pandas as pd
import os, uuid, json
from datetime import date

CLAN_OPTIONS = ["AlterEgo", "Nirvana", "AE Hells"]

main_bp = Blueprint('main', __name__)


def parse_txt_to_df(filepath):
    for enc in ('utf-8', 'cp1250', 'latin-1'):
        try:
            df = pd.read_csv(filepath, encoding=enc, encoding_errors='replace')
            return df
        except Exception:
            continue
    raise ValueError("Could not parse file with any known encoding.")


def t():
    return get_translations(session.get('lang', 'en'))


@main_bp.route('/')
@login_required
def dashboard():
    my_uploads = Upload.query.filter_by(user_id=current_user.id).order_by(Upload.uploaded_at.desc()).all()
    shared_uploads = Upload.query.filter_by(is_shared=True).order_by(Upload.uploaded_at.desc()).all()
    my_results = Result.query.filter_by(user_id=current_user.id).order_by(Result.created_at.desc()).limit(5).all()
    return render_template('dashboard.html',
                           my_uploads=my_uploads,
                           shared_uploads=shared_uploads,
                           my_results=my_results)


@main_bp.route('/upload', methods=['GET', 'POST'])
@login_required
def upload():
    tr = t()
    if request.method == 'POST':
        file = request.files.get('csv_file')
        clan = request.form.get('clan', '').strip()

        if not file or file.filename == '':
            flash(tr['upload_error_no_file'], 'error')
            return redirect(request.url)
        if not file.filename.lower().endswith('.txt'):
            flash(tr['upload_error_wrong_type'], 'error')
            return redirect(request.url)
        if clan not in CLAN_OPTIONS:
            flash(tr['upload_error_no_clan'], 'error')
            return redirect(request.url)

        # Internal storage filename (uuid to avoid collisions)
        filename = f"{uuid.uuid4().hex}.txt"
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        try:
            df = parse_txt_to_df(filepath)
            row_count = len(df)
            columns = json.dumps(list(df.columns))
        except Exception as e:
            os.remove(filepath)
            flash(f'Could not parse file: {e}', 'error')
            return redirect(request.url)

        # Display name: ClanName_YYYY-MM-DD.txt
        today = date.today().strftime('%Y-%m-%d')
        display_name = f"{clan}_{today}.txt"

        is_shared = 'is_shared' in request.form and current_user.is_admin()

        upload_rec = Upload(
            filename=filename,
            original_filename=display_name,
            user_id=current_user.id,
            is_shared=is_shared,
            row_count=row_count,
            columns=columns,
        )
        db.session.add(upload_rec)
        db.session.commit()
        flash(f'"{display_name}" {tr["upload_success"]} — {row_count} {tr["upload_rows_found"]}.', 'success')
        return redirect(url_for('main.view_upload', upload_id=upload_rec.id))

    return render_template('upload.html', clan_options=CLAN_OPTIONS)


@main_bp.route('/uploads')
@login_required
def manage_uploads():
    if current_user.is_admin():
        all_uploads = Upload.query.order_by(Upload.uploaded_at.desc()).all()
    else:
        all_uploads = Upload.query.filter(
            (Upload.user_id == current_user.id) | (Upload.is_shared == True)
        ).order_by(Upload.uploaded_at.desc()).all()
    return render_template('manage_uploads.html', uploads=all_uploads)


@main_bp.route('/uploads/<int:upload_id>/preview')
@login_required
def preview_upload(upload_id):
    tr = t()
    upload_rec = _get_accessible_upload(upload_id)
    if not upload_rec:
        flash(tr['flash_access_denied'], 'error')
        return redirect(url_for('main.manage_uploads'))
    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], upload_rec.filename)
    try:
        df = parse_txt_to_df(filepath)
        columns = list(df.columns)
        rows = df.head(50).fillna('').to_dict(orient='records')
        total_rows = len(df)
        dtypes = {col: str(df[col].dtype) for col in columns}
        stats = {}
        for col in df.select_dtypes(include='number').columns:
            stats[col] = {
                'min': round(df[col].min(), 2),
                'max': round(df[col].max(), 2),
                'mean': round(df[col].mean(), 2),
                'nulls': int(df[col].isnull().sum()),
            }
    except Exception as e:
        flash(f'Could not read file: {e}', 'error')
        return redirect(url_for('main.manage_uploads'))
    return render_template('preview_upload.html',
                           upload=upload_rec,
                           columns=columns,
                           rows=rows,
                           total_rows=total_rows,
                           dtypes=dtypes,
                           stats=stats)


@main_bp.route('/choices/<int:upload_id>', methods=['GET', 'POST'])
@login_required
def choices(upload_id):
    tr = t()
    upload_rec = _get_accessible_upload(upload_id)
    if not upload_rec:
        flash(tr['choices_error'], 'error')
        return redirect(url_for('main.dashboard'))

    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], upload_rec.filename)
    df = parse_txt_to_df(filepath)
    choice_defs = get_choice_options(df)

    if request.method == 'POST':
        user_choices = {}
        for c in choice_defs:
            if c['type'] == 'checkbox':
                user_choices[c['name']] = c['name'] in request.form
            else:
                user_choices[c['name']] = request.form.get(c['name'], c.get('default', ''))

        try:
            output = process(df, user_choices)
        except Exception as e:
            flash(f'{tr["choices_proc_error"]}: {e}', 'error')
            return redirect(request.url)

        result = Result(
            upload_id=upload_rec.id,
            user_id=current_user.id,
            choices=json.dumps(user_choices),
            output=json.dumps(output),
        )
        db.session.add(result)
        db.session.commit()
        return redirect(url_for('main.results', result_id=result.id))

    return render_template('choices.html', upload=upload_rec, choice_defs=choice_defs)


@main_bp.route('/results/<int:result_id>')
@login_required
def results(result_id):
    tr = t()
    result = Result.query.get_or_404(result_id)
    if result.user_id != current_user.id and not current_user.is_admin():
        flash(tr['flash_access_denied'], 'error')
        return redirect(url_for('main.dashboard'))
    output = json.loads(result.output)
    choices = json.loads(result.choices)
    return render_template('results.html', result=result, output=output, choices=choices)


@main_bp.route('/uploads/<int:upload_id>/delete', methods=['POST'])
@login_required
def delete_upload(upload_id):
    tr = t()
    upload_rec = Upload.query.get_or_404(upload_id)
    if upload_rec.user_id != current_user.id and not current_user.is_admin():
        flash(tr['flash_access_denied'], 'error')
        return redirect(url_for('main.manage_uploads'))
    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], upload_rec.filename)
    if os.path.exists(filepath):
        os.remove(filepath)
    db.session.delete(upload_rec)
    db.session.commit()
    flash(tr['flash_upload_deleted'], 'success')
    return redirect(url_for('main.manage_uploads'))


def _get_accessible_upload(upload_id):
    upload_rec = Upload.query.get(upload_id)
    if not upload_rec:
        return None
    if upload_rec.user_id == current_user.id:
        return upload_rec
    if upload_rec.is_shared:
        return upload_rec
    if current_user.is_admin():
        return upload_rec
    return None


@main_bp.route('/uploads/<int:upload_id>/view')
@login_required
def view_upload(upload_id):
    upload_rec = _get_accessible_upload(upload_id)
    if not upload_rec:
        flash(t()['flash_access_denied'], 'error')
        return redirect(url_for('main.manage_uploads'))
    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], upload_rec.filename)
    try:
        df = parse_txt_to_df(filepath)
        columns = list(df.columns)
        rows = df.fillna('').to_dict(orient='records')
    except Exception as e:
        flash(f'Could not read file: {e}', 'error')
        return redirect(url_for('main.manage_uploads'))

    # Load existing attendance for this upload
    attendance = Attendance.query.filter_by(upload_id=upload_id).first()
    confirmed = json.loads(attendance.confirmed_rows) if attendance else {}
    power     = json.loads(attendance.power_rows  if attendance and attendance.power_rows  else '{}')
    absent    = json.loads(attendance.absent_rows if attendance and attendance.absent_rows else '{}')

    confirmed_count = sum(1 for v in confirmed.values() if v)
    power_count     = sum(1 for v in power.values() if v)
    absent_count    = sum(1 for v in absent.values() if v)

    return render_template('view_upload.html',
                           upload=upload_rec,
                           columns=columns,
                           rows=rows,
                           confirmed=confirmed,
                           power=power,
                           absent=absent,
                           confirmed_count=confirmed_count,
                           power_count=power_count,
                           absent_count=absent_count)


@main_bp.route('/uploads/<int:upload_id>/attendance', methods=['POST'])
@login_required
def save_attendance(upload_id):
    upload_rec = _get_accessible_upload(upload_id)
    if not upload_rec:
        flash(t()['flash_access_denied'], 'error')
        return redirect(url_for('main.manage_uploads'))

    # Collect checked row indices from form
    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], upload_rec.filename)
    df = parse_txt_to_df(filepath)
    confirmed = {}
    power     = {}
    absent    = {}
    for i in range(len(df)):
        confirmed[str(i)] = request.form.get(f'row_{i}',    '0') == '1'
        power[str(i)]     = request.form.get(f'power_{i}',  '0') == '1'
        absent[str(i)]    = request.form.get(f'absent_{i}', '0') == '1'

    attendance = Attendance.query.filter_by(upload_id=upload_id).first()
    if attendance:
        attendance.confirmed_rows = json.dumps(confirmed)
        attendance.power_rows     = json.dumps(power)
        attendance.absent_rows    = json.dumps(absent)
        attendance.saved_at = __import__('datetime').datetime.utcnow()
    else:
        attendance = Attendance(upload_id=upload_id,
                                confirmed_rows=json.dumps(confirmed),
                                power_rows=json.dumps(power),
                                absent_rows=json.dumps(absent))
        db.session.add(attendance)
    db.session.commit()

    tr = t()
    flash(tr['power_player_saved'], 'success')
    return redirect(url_for('main.view_upload', upload_id=upload_id))


# ---------------------------------------------------------------------------
# ROSTER GENERATOR
# ---------------------------------------------------------------------------

BATTLE_TYPES        = ['RoE', 'Clan Battle']
CLAN_BATTLE_MODES   = ['Standard', '8 4 2 1']
PRIORITY_OPTIONS    = ['Rezonowanie', 'Class']
ONLINE_OPTIONS      = ['Only confirmed online', 'Prioritize Online', 'All players', 'Exclude Absent']
DIST_OPTIONS        = ['Max power', 'Even distribution']
ROE_BATTLES_OPTIONS = ['7', '10']
POWER_PLAYER_OPTIONS = ['Yes — apply 20% boost', 'No boost']

ALL_CLASSES = [
    'Nekromanta', 'Barbarzyńca', 'Łowca demonów', 'Rycerz krwi',
    'Czarownik', 'Mnich', 'Krzyżowiec', 'Sztormiciel', 'Druid'
]


# ---------------------------------------------------------------------------
# Grouping helpers
# ---------------------------------------------------------------------------

def _reso(p):
    try:
        return int(p.get('Rezonowanie', 0) or 0)
    except Exception:
        return 0


def _sort_by_priority(players, priority):
    """Return players sorted according to priority setting."""
    if priority == 'Rezonowanie':
        return sorted(players, key=_reso, reverse=True)
    else:
        # Class round-robin (sorted by reso within class)
        from collections import defaultdict
        by_class = defaultdict(list)
        for p in sorted(players, key=_reso, reverse=True):
            by_class[p.get('Klasa', '')].append(p)
        result = []
        classes = sorted(by_class.keys(), key=lambda c: -len(by_class[c]))
        while any(by_class[c] for c in classes):
            for c in classes:
                if by_class[c]:
                    result.append(by_class[c].pop(0))
        return result


def _distribute(pool, num_groups, group_size, distribution):
    """Split pool into num_groups of group_size using chosen distribution."""
    if distribution == 'Max power':
        groups = []
        for i in range(num_groups):
            groups.append(pool[i*group_size:(i+1)*group_size])
        return groups
    else:
        # Snake draft for even reso averages
        by_reso = sorted(pool, key=_reso, reverse=True)
        groups = [[] for _ in range(num_groups)]
        for i, player in enumerate(by_reso):
            rnd = i // num_groups
            pos = i % num_groups
            idx = pos if rnd % 2 == 0 else num_groups - 1 - pos
            groups[idx].append(player)
        return groups


def _build_groups(players, num_battles, priority, distribution,
                  clan_mode='Standard'):
    """
    Main grouping entry point.
    clan_mode: 'Standard' | '8 4 2 1'
    """
    group_size = 8

    # Enrich with numeric reso
    for p in players:
        p['_reso'] = _reso(p)

    sorted_pool = _sort_by_priority(players, priority)

    if clan_mode == '8 4 2 1':
        # Tier sizes: 3 groups of 8 = 24 players per tier label
        # Label tiers: first 3 groups → "8", next 3 → "4", next 3 → "2", last 3 → "1"
        # Total: 12 groups of 8 = 96 slots
        tier_labels = ['8']*3 + ['4']*3 + ['2']*3 + ['1']*3
        total_slots = 12 * group_size
        pool = sorted_pool[:total_slots]

        tier_groups = []
        for tier_start in range(0, 12, 3):
            tier_pool = pool[tier_start*group_size:(tier_start+3)*group_size]
            tier_g = _distribute(tier_pool, 3, group_size, distribution)
            for g in tier_g:
                tier_groups.append(g)

        for p in players:
            p.pop('_reso', None)
        return tier_groups, tier_labels

    else:
        total_slots = num_battles * group_size
        pool = sorted_pool[:total_slots]
        groups = _distribute(pool, num_battles, group_size, distribution)
        for p in players:
            p.pop('_reso', None)
        return groups, None


def _apply_online_filter(all_players, online_mode, num_slots):
    """
    online_mode may be a string (legacy) or a list of option strings.
    Options can be combined; logic:
      'Only confirmed online'  → include only _confirmed players
      'Prioritize Online'      → confirmed first, then fill with remaining
      'Exclude Absent'         → remove _absent players (applied as a post-filter)
      'All players'            → no pre-filter (default if nothing else chosen)
    When multiple are selected, they compose:
      e.g. [Prioritize Online, Exclude Absent] = prioritize confirmed,
           but also strip absent from the pool first.
    """
    # Normalise to list
    if isinstance(online_mode, str):
        modes = [online_mode]
    else:
        modes = list(online_mode) if online_mode else ['All players']

    pool = list(all_players)

    # Step 1: exclude absent if requested
    if 'Exclude Absent' in modes:
        pool = [p for p in pool if not p.get('_absent')]

    # Step 2: apply presence filter
    if 'Only confirmed online' in modes:
        pool = [p for p in pool if p.get('_confirmed')]
    elif 'Prioritize Online' in modes:
        confirmed = [p for p in pool if p.get('_confirmed')]
        remaining = [p for p in pool if not p.get('_confirmed')]
        pool = (confirmed + remaining)[:num_slots]
    # 'All players' or no presence filter → keep pool as-is

    return pool


def _avg_reso(group):
    vals = [_reso(p) for p in group]
    return round(sum(vals)/len(vals)) if vals else 0


# ---------------------------------------------------------------------------
# Roster routes
# ---------------------------------------------------------------------------

@main_bp.route('/roster', methods=['GET', 'POST'])
@login_required
def roster_select():
    tr = t()
    if current_user.is_admin():
        all_uploads = Upload.query.order_by(Upload.uploaded_at.desc()).all()
    else:
        all_uploads = Upload.query.filter(
            (Upload.user_id == current_user.id) | (Upload.is_shared == True)
        ).order_by(Upload.uploaded_at.desc()).all()

    if request.method == 'POST':
        selected = request.form.getlist('upload_ids')
        if not selected:
            flash(tr['roster_error_no_files'], 'error')
            return redirect(request.url)
        if len(selected) > 3:
            flash(tr['roster_error_too_many'], 'error')
            return redirect(request.url)
        ids_str = ','.join(selected[:3])
        return redirect(url_for('main.roster_config', ids=ids_str))

    return render_template('roster_select.html', uploads=all_uploads)


@main_bp.route('/roster/config', methods=['GET', 'POST'])
@login_required
def roster_config():
    tr = t()
    ids_str = request.args.get('ids', '') or request.form.get('ids', '')
    try:
        upload_ids = [int(i) for i in ids_str.split(',') if i.strip()]
    except ValueError:
        flash(tr['roster_error_no_files'], 'error')
        return redirect(url_for('main.roster_select'))

    uploads = [_get_accessible_upload(uid) for uid in upload_ids]
    uploads = [u for u in uploads if u]
    if not uploads:
        flash(tr['roster_error_no_files'], 'error')
        return redirect(url_for('main.roster_select'))

    def _render_config(**kw):
        # Build value->label map using current language
        tr = t()
        opt_labels = {
            'RoE':                       tr['opt_roe'],
            'Clan Battle':               tr['opt_clan_battle'],
            'Standard':                  tr['opt_standard'],
            '8 4 2 1':                   tr['opt_8421'],
            'Rezonowanie':               tr['opt_reso'],
            'Class':                     tr['opt_class'],
            'Only confirmed online':     tr['opt_online_only'],
            'Prioritize Online':         tr['opt_prioritize_online'],
            'All players':               tr['opt_all_players'],
            'Max power':                 tr['opt_max_power'],
            'Even distribution':         tr['opt_even_dist'],
            '7':                         tr['opt_7'],
            '10':                        tr['opt_10'],
            'Yes — apply 20% boost':     tr['opt_power_yes'],
            'No boost':                  tr['opt_power_no'],
            'Exclude Absent':            tr['opt_exclude_absent'],
        }
        return render_template('roster_config.html',
                               uploads=uploads, ids_str=ids_str,
                               battle_types=BATTLE_TYPES,
                               clan_battle_modes=CLAN_BATTLE_MODES,
                               priority_options=PRIORITY_OPTIONS,
                               online_options=ONLINE_OPTIONS,
                               dist_options=DIST_OPTIONS,
                               roe_battles_options=ROE_BATTLES_OPTIONS,
                               power_player_options=POWER_PLAYER_OPTIONS,
                               all_classes=ALL_CLASSES,
                               opt_labels=opt_labels,
                               **kw)

    if request.method == 'POST':
        battle_type   = request.form.get('battle_type', '').strip()
        clan_mode     = request.form.get('clan_mode', 'Standard').strip()
        priority      = request.form.get('priority', '').strip()
        online_only   = request.form.getlist('online_only') or ['All players']
        distribution  = request.form.get('distribution', '').strip()
        roe_battles   = request.form.get('roe_battles', '10').strip()
        active_classes= request.form.getlist('active_classes')
        power_player  = request.form.get('power_player', '').strip()

        errors = []
        if battle_type not in BATTLE_TYPES:
            errors.append(tr['roster_error_no_battle_type'])
        if priority not in PRIORITY_OPTIONS:
            errors.append(tr['roster_error_no_priority'])
        if not all(o in ONLINE_OPTIONS for o in online_only):
            errors.append(tr['roster_error_no_online'])
        if distribution not in DIST_OPTIONS:
            errors.append(tr['roster_error_no_dist'])
        if power_player not in POWER_PLAYER_OPTIONS:
            errors.append(tr['roster_error_no_power'])
        if not active_classes:
            errors.append(tr['roster_error_no_classes'])

        if errors:
            for e in errors:
                flash(e, 'error')
            return _render_config()

        if battle_type == 'RoE':
            num_battles = int(roe_battles) if roe_battles in ROE_BATTLES_OPTIONS else 10
        elif clan_mode == '8 4 2 1':
            num_battles = 12
        else:
            num_battles = 12

        session['roster_ids']         = upload_ids
        session['roster_battle_type'] = battle_type
        session['roster_clan_mode']   = clan_mode
        session['roster_priority']    = priority
        session['roster_online_only'] = json.dumps(online_only)
        session['roster_distribution']= distribution
        session['roster_num_battles'] = num_battles
        session['roster_active_classes'] = active_classes
        session['roster_power_player'] = power_player
        return redirect(url_for('main.roster_view'))

    return _render_config()


@main_bp.route('/roster/view')
@login_required
def roster_view():
    tr = t()
    upload_ids     = session.get('roster_ids', [])
    battle_type    = session.get('roster_battle_type', '')
    clan_mode      = session.get('roster_clan_mode', 'Standard')
    priority       = session.get('roster_priority', 'Rezonowanie')
    online_only_raw = session.get('roster_online_only', '["All players"]')
    try:
        online_only = json.loads(online_only_raw) if isinstance(online_only_raw, str) and online_only_raw.startswith('[') else [online_only_raw]
    except Exception:
        online_only = ['All players']
    distribution   = session.get('roster_distribution', 'Max power')
    num_battles    = session.get('roster_num_battles', 12)
    active_classes = session.get('roster_active_classes', ALL_CLASSES)
    power_player   = session.get('roster_power_player', 'No boost')

    if not upload_ids or not battle_type:
        flash(tr['roster_error_no_files'], 'error')
        return redirect(url_for('main.roster_select'))

    all_players = []
    source_files = []
    columns_used = None

    for uid in upload_ids:
        upload_rec = _get_accessible_upload(uid)
        if not upload_rec:
            continue
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], upload_rec.filename)
        try:
            df = parse_txt_to_df(filepath)
        except Exception:
            continue

        attendance = Attendance.query.filter_by(upload_id=uid).first()
        confirmed_map = json.loads(attendance.confirmed_rows) if attendance else {}

        cols = list(df.columns)
        if columns_used is None:
            columns_used = cols[1:-1]

        power_map  = json.loads(attendance.power_rows  if attendance and attendance.power_rows  else '{}')
        absent_map = json.loads(attendance.absent_rows if attendance and attendance.absent_rows else '{}')
        rows = df.fillna('').to_dict(orient='records')
        for i, row in enumerate(rows):
            row['_confirmed'] = confirmed_map.get(str(i), False)
            row['_power']     = power_map.get(str(i), False)
            row['_absent']    = absent_map.get(str(i), False)
            row['_source']    = upload_rec.original_filename
            all_players.append(row)
        source_files.append(upload_rec.original_filename)

    # Deduplicate by name
    seen = {}
    for p in sorted(all_players, key=_reso, reverse=True):
        name = p.get('Nazwa', '')
        if name and name not in seen:
            seen[name] = p
    all_players = list(seen.values())

    # Class filter
    all_players = [p for p in all_players
                   if p.get('Klasa', '') in active_classes]

    # PowerPlayer boost — mark players who are power_player and inflate their reso
    if power_player == 'Yes — apply 20% boost':
        for p in all_players:
            if p.get('_power', False):
                try:
                    original = int(p.get('Rezonowanie', 0) or 0)
                    p['Rezonowanie'] = str(round(original * 1.20))
                    p['_boosted'] = True
                except Exception:
                    pass

    # Online filter
    total_slots = num_battles * 8
    pool = _apply_online_filter(all_players, online_only, total_slots)

    # Build groups
    groups, tier_labels = _build_groups(
        pool, num_battles, priority, distribution,
        clan_mode=clan_mode if battle_type == 'Clan Battle' else 'Standard'
    )

    group_stats = [{'avg_reso': _avg_reso(g), 'count': len(g)} for g in groups]

    # Strip internal _ keys from groups and pool before passing to template
    # Rename _boosted -> boosted, _source -> source_file so they survive the underscore-strip
    for g in groups:
        for p in g:
            if p.get('_boosted'):
                p['boosted'] = True
            if p.get('_source'):
                p['source_file'] = p['_source']
    for p in all_players:
        if p.get('_source'):
            p['source_file'] = p['_source']

    def _clean(p):
        return {k: v for k, v in p.items() if not k.startswith('_')}

    clean_groups = [[_clean(p) for p in g] for g in groups]
    full_pool_clean = [_clean(p) for p in all_players]

    # Store roster data server-side in a temp file (session too small for large rosters)
    import tempfile, pickle
    roster_tmp = {
        'groups':       clean_groups,
        'columns':      columns_used or [],
        'player_pool':  full_pool_clean,
        'tier_labels':  tier_labels,
        'source_files': source_files,
    }
    tmp_path = os.path.join(current_app.config['UPLOAD_FOLDER'], f'_roster_tmp_{current_user.id}.pkl')
    with open(tmp_path, 'wb') as f:
        pickle.dump(roster_tmp, f)

    tr = t()
    opt_labels = {
        'RoE': tr['opt_roe'], 'Clan Battle': tr['opt_clan_battle'],
        'Standard': tr['opt_standard'], '8 4 2 1': tr['opt_8421'],
        'Rezonowanie': tr['opt_reso'], 'Class': tr['opt_class'],
        'Only confirmed online': tr['opt_online_only'],
        'Prioritize Online': tr['opt_prioritize_online'],
        'All players': tr['opt_all_players'],
        'Exclude Absent': tr['opt_exclude_absent'],
        'Max power': tr['opt_max_power'],
        'Even distribution': tr['opt_even_dist'],
    }
    source_color_map = {sf: i for i, sf in enumerate(source_files)}
    _ol_map = {'Only confirmed online': tr['opt_online_only'], 'Prioritize Online': tr['opt_prioritize_online'],
               'All players': tr['opt_all_players'], 'Exclude Absent': tr['opt_exclude_absent']}
    online_label = ' + '.join(_ol_map.get(o, o) for o in (online_only if isinstance(online_only, list) else [online_only]))

    # Build list of absent players for display always
    absent_players = [_clean(p) for p in all_players if p.get('_absent')]

    return render_template('roster_view.html',
                           groups=clean_groups,
                           group_stats=group_stats,
                           tier_labels=tier_labels,
                           columns=columns_used or [],
                           battle_type=battle_type,
                           clan_mode=clan_mode,
                           priority=priority,
                           online_only=online_only,
                           online_label=online_label,
                           distribution=distribution,
                           num_battles=num_battles,
                           active_classes=active_classes,
                           source_files=source_files,
                           total_players=len(pool),
                           full_pool=full_pool_clean,
                           power_player=power_player,
                           opt_labels=opt_labels,
                           source_color_map=source_color_map,
                           absent_players=absent_players)


@main_bp.route('/roster/save', methods=['POST'])
@login_required
def roster_save():
    from models import SavedRoster
    import traceback
    tr = t()

    import pickle
    tmp_path = os.path.join(current_app.config['UPLOAD_FOLDER'], f'_roster_tmp_{current_user.id}.pkl')
    if not os.path.exists(tmp_path):
        flash('Session expired — please regenerate the roster first.', 'error')
        return redirect(url_for('main.roster_select'))
    with open(tmp_path, 'rb') as f:
        save_data = pickle.load(f)

    battle_type = session.get('roster_battle_type', '')
    clan_mode   = session.get('roster_clan_mode', 'Standard')

    config = {
        'battle_type':    battle_type,
        'clan_mode':      clan_mode,
        'priority':       session.get('roster_priority'),
        'online_only':    online_only,
        'distribution':   session.get('roster_distribution'),
        'num_battles':    session.get('roster_num_battles'),
        'active_classes': session.get('roster_active_classes'),
        'source_files':   save_data.get('source_files', []),
        'tier_labels':    save_data.get('tier_labels'),
    }

    today = date.today().strftime('%Y-%m-%d')
    name  = f"{current_user.username}_{today}"

    try:
        sr = SavedRoster(
            name=name,
            created_by=current_user.id,
            battle_type=battle_type,
            config=json.dumps(config),
            groups_data=json.dumps(save_data.get('groups', [])),
            columns_data=json.dumps(save_data.get('columns', [])),
            player_pool=json.dumps(save_data.get('player_pool', [])),
            overrides='{}',
        )
        db.session.add(sr)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'roster_save error: {traceback.format_exc()}')
        flash(f'Could not save roster: {e}', 'error')
        return redirect(url_for('main.roster_view'))

    flash(f'{tr["roster_saved"]} "{name}"', 'success')
    return redirect(url_for('main.saved_roster_view', roster_id=sr.id))


# ---------------------------------------------------------------------------
# Saved Rosters module
# ---------------------------------------------------------------------------

@main_bp.route('/saved-rosters')
@login_required
def saved_rosters_list():
    from models import SavedRoster
    if current_user.is_admin():
        rosters = SavedRoster.query.order_by(SavedRoster.created_at.desc()).all()
    else:
        rosters = SavedRoster.query.filter_by(created_by=current_user.id)\
                             .order_by(SavedRoster.created_at.desc()).all()
    return render_template('saved_rosters_list.html', rosters=rosters)


@main_bp.route('/saved-rosters/<int:roster_id>')
@login_required
def saved_roster_view(roster_id):
    from models import SavedRoster
    sr = SavedRoster.query.get_or_404(roster_id)
    if sr.created_by != current_user.id and not current_user.is_admin():
        flash(t()['flash_access_denied'], 'error')
        return redirect(url_for('main.saved_rosters_list'))

    groups      = json.loads(sr.groups_data)
    columns     = json.loads(sr.columns_data)
    player_pool = json.loads(sr.player_pool)
    config      = json.loads(sr.config)
    tier_labels = config.get('tier_labels')

    # groups_data already contains the current display state (overrides baked in)
    group_stats = [{'avg_reso': _avg_reso(g), 'count': len(g)} for g in groups]

    sf = config.get('source_files', [])
    if isinstance(sf, str):
        try: sf = json.loads(sf)
        except: sf = []
    source_color_map = {f: i for i, f in enumerate(sf)}

    # Extract absent players from pool for display
    absent_players = [p for p in player_pool if p.get('absent')]

    return render_template('saved_roster_view.html',
                           sr=sr,
                           groups=groups,
                           group_stats=group_stats,
                           tier_labels=tier_labels,
                           columns=columns,
                           config=config,
                           player_pool=player_pool,
                           source_files=sf,
                           source_color_map=source_color_map,
                           absent_players=absent_players)


@main_bp.route('/saved-rosters/<int:roster_id>/override', methods=['POST'])
@login_required
def saved_roster_override(roster_id):
    from models import SavedRoster
    sr = SavedRoster.query.get_or_404(roster_id)
    if sr.created_by != current_user.id and not current_user.is_admin():
        flash(t()['flash_access_denied'], 'error')
        return redirect(url_for('main.saved_rosters_list'))

    # The form submits the full current state as JSON via hidden field 'roster_state'
    roster_state_json = request.form.get('roster_state', '')
    if roster_state_json:
        try:
            # roster_state: list of groups, each group = list of player dicts (already in display order)
            new_groups = json.loads(roster_state_json)
            sr.groups_data = json.dumps(new_groups)
            sr.overrides   = '{}'   # overrides are now baked into groups_data
            db.session.commit()
            flash(t()['roster_overrides_saved'], 'success')
        except Exception as e:
            import traceback
            current_app.logger.error(f'override save error: {traceback.format_exc()}')
            flash(f'Save failed: {e}', 'error')
    else:
        flash('No data received.', 'error')

    return redirect(url_for('main.saved_roster_view', roster_id=roster_id))


@main_bp.route('/saved-rosters/<int:roster_id>/delete', methods=['POST'])
@login_required
def saved_roster_delete(roster_id):
    from models import SavedRoster
    sr = SavedRoster.query.get_or_404(roster_id)
    if sr.created_by != current_user.id and not current_user.is_admin():
        flash(t()['flash_access_denied'], 'error')
        return redirect(url_for('main.saved_rosters_list'))
    db.session.delete(sr)
    db.session.commit()
    flash(t()['roster_deleted'], 'success')
    return redirect(url_for('main.saved_rosters_list'))




@main_bp.route('/saved-rosters/<int:roster_id>/export')
@login_required
def saved_roster_export(roster_id):
    import traceback
    from models import SavedRoster
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    sr = SavedRoster.query.get_or_404(roster_id)
    if sr.created_by != current_user.id and not current_user.is_admin():
        flash(t()['flash_access_denied'], 'error')
        return redirect(url_for('main.saved_rosters_list'))

    try:
        groups    = json.loads(sr.groups_data  or '[]')
        columns   = json.loads(sr.columns_data or '[]')
        overrides = json.loads(sr.overrides    or '{}')
        config    = json.loads(sr.config       or '{}')
    except Exception as e:
        flash(f'Could not read roster data: {e}', 'error')
        return redirect(url_for('main.saved_roster_view', roster_id=roster_id))

    if not groups or not columns:
        flash('Roster has no data — please regenerate and save it again.', 'error')
        return redirect(url_for('main.saved_roster_view', roster_id=roster_id))

    tier_labels = config.get('tier_labels')

    # Apply overrides
    display_groups = []
    for gi, group in enumerate(groups):
        dg = []
        for si, player in enumerate(group):
            key = f'{gi}:{si}'
            dg.append(overrides.get(key, player))
        display_groups.append(dg)

    try:
        # ── Colours ────────────────────────────────────────────────────
        black      = '000000'
        white      = 'FFFFFF'
        light_gray = 'F2F2F2'
        mid_gray   = 'CCCCCC'
        dark_gray  = '444444'
        hdr_fill   = '2A2A2A'
        hdr_text   = 'FFFFFF'
        accent_dk  = '3A6B00'

        tier_colours = {
            '8': ('FFF0E0', '7A3A00'),
            '4': ('E8F5E8', '1A5C1A'),
            '2': ('E0EEF8', '0A3A6A'),
            '1': ('FCE8E8', '7A1A1A'),
        }
        class_colours = {
            'Nekromanta':    ('EEF0FF', '3A3A8A'),
            'Barbarzyca':    ('FFF0E0', '7A3A00'),
            'Barbarzyńca':   ('FFF0E0', '7A3A00'),
            'Łowca demonów': ('EAFAEA', '1A5C1A'),
            'Rycerz krwi':   ('FFE8E8', '7A1A1A'),
            'Czarownik':     ('F5E8FF', '6A1A8A'),
            'Mnich':         ('E8F8FF', '0A5A7A'),
            'Krzyżowiec':    ('FFFAE0', '6A5A00'),
            'Sztormiciel':   ('E8F0FF', '1A3A8A'),
            'Druid':         ('E8FFEE', '0A5A2A'),
        }

        def sol(hex_col):
            return PatternFill('solid', start_color=hex_col, fgColor=hex_col)

        def bdr(color=mid_gray):
            s = Side(style='thin', color=color)
            return Border(left=s, right=s, top=s, bottom=s)

        # ── Layout constants ──────────────────────────────────────────
        BATTLES_PER_ROW = 3
        GAP_COLS        = 1   # empty spacer column between battles
        # Each battle block occupies: 1 (row#) + len(columns) cols
        BATTLE_COLS = 1 + len(columns)

        wb = Workbook()
        ws = wb.active
        ws.title = sr.name[:31]
        ws.sheet_view.showGridLines = False

        # ── Helper: write one battle block ────────────────────────────
        def write_battle(gi, group, start_row, start_col):
            tier = tier_labels[gi] if tier_labels and gi < len(tier_labels) else None
            if tier and tier_labels:
                same_tier_before = sum(1 for t in tier_labels[:gi] if t == tier)
                pos_in_tier = same_tier_before + 1
                battle_label = f'Battle {gi+1} ({pos_in_tier} for {tier}pts)'
            else:
                battle_label = f'Battle {gi+1}'
            tier_fill, tier_text = (tier_colours.get(tier, (light_gray, dark_gray))
                                    if tier else (light_gray, dark_gray))

            # Battle header row
            ws.row_dimensions[start_row].height = 18
            end_col = start_col + BATTLE_COLS - 1
            c = ws.cell(start_row, start_col, battle_label)
            c.font      = Font(name='Arial', bold=True, size=9, color=tier_text)
            c.fill      = sol(tier_fill)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = bdr()
            ws.merge_cells(start_row=start_row, start_column=start_col,
                           end_row=start_row, end_column=end_col)

            # Column sub-header row
            sub_row = start_row + 1
            ws.row_dimensions[sub_row].height = 16
            sub_cols = ['#'] + columns
            for ci_offset, hdr in enumerate(sub_cols):
                c = ws.cell(sub_row, start_col + ci_offset, hdr)
                c.font      = Font(name='Arial', bold=True, size=8, color=hdr_text)
                c.fill      = sol(hdr_fill)
                c.alignment = Alignment(
                    horizontal='center' if ci_offset == 0 else 'left',
                    vertical='center')
                c.border    = bdr(hdr_fill)

            # Player rows
            for pi, player in enumerate(group):
                data_row = start_row + 2 + pi
                ws.row_dimensions[data_row].height = 15
                rbg = white if pi % 2 == 0 else light_gray

                # Row number
                c = ws.cell(data_row, start_col, pi + 1)
                c.font      = Font(name='Arial', size=8, color=dark_gray)
                c.fill      = sol(rbg)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border    = bdr()

                for ci_offset, col in enumerate(columns, 1):
                    val = str(player.get(col, '') or '')
                    c = ws.cell(data_row, start_col + ci_offset, val)
                    c.fill      = sol(rbg)
                    c.alignment = Alignment(horizontal='left', vertical='center')
                    c.font      = Font(name='Arial', size=8, color=black)
                    c.border    = bdr()

                    if col == 'Nazwa':
                        c.font = Font(name='Arial', size=8, bold=True, color=black)
                    elif col == 'Klasa':
                        bg, fg = class_colours.get(val, (rbg, dark_gray))
                        c.fill = sol(bg)
                        c.font = Font(name='Arial', size=8, bold=True, color=fg)
                        c.alignment = Alignment(horizontal='center', vertical='center')
                    elif col in ('Rezonowanie', 'Ranking udziału', 'Poziom'):
                        c.alignment = Alignment(horizontal='center', vertical='center')
                        if val in ('Poza rankingiem', ''):
                            c.font = Font(name='Arial', size=8, color=mid_gray,
                                          italic=True)

            # Avg reso footer row
            try:
                avg = round(sum(int(p.get('Rezonowanie', 0) or 0)
                                for p in group) / len(group)) if group else 0
            except Exception:
                avg = 0
            footer_row = start_row + 2 + len(group)
            ws.row_dimensions[footer_row].height = 12
            c = ws.cell(footer_row, start_col,
                        f'{len(group)} players  ·  avg reso {avg}')
            c.font      = Font(name='Arial', size=7, color=dark_gray, italic=True)
            c.fill      = sol(light_gray)
            c.alignment = Alignment(horizontal='left', vertical='center')
            c.border    = bdr()
            ws.merge_cells(start_row=footer_row, start_column=start_col,
                           end_row=footer_row, end_column=end_col)

            # Return how many rows this block consumed
            return 2 + len(group) + 1  # header + sub-header + players + footer

        # ── Title + meta (row 1-2) ────────────────────────────────────
        total_cols = BATTLES_PER_ROW * BATTLE_COLS + (BATTLES_PER_ROW - 1) * GAP_COLS

        ws.row_dimensions[1].height = 22
        c = ws.cell(1, 1, sr.name)
        c.font      = Font(name='Arial', bold=True, size=13, color=accent_dk)
        c.fill      = sol(light_gray)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=total_cols)

        ws.row_dimensions[2].height = 13
        meta_parts = [p for p in [
            sr.battle_type,
            config.get('priority'),
            config.get('online_only'),
            config.get('distribution'),
        ] if p]
        c = ws.cell(2, 1, '  ·  '.join(meta_parts))
        c.font      = Font(name='Arial', size=8, color=dark_gray, italic=True)
        c.fill      = sol(light_gray)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2, end_column=total_cols)

        # ── Set column widths ─────────────────────────────────────────
        # Repeat the same width pattern for each of the 3 battle columns
        col_widths = [4] + [max(13, len(c) + 2) for c in columns]
        for block in range(BATTLES_PER_ROW):
            base_col = 1 + block * (BATTLE_COLS + GAP_COLS)
            for ci_offset, w in enumerate(col_widths):
                ws.column_dimensions[
                    get_column_letter(base_col + ci_offset)].width = w
            # spacer column
            if block < BATTLES_PER_ROW - 1:
                ws.column_dimensions[
                    get_column_letter(base_col + BATTLE_COLS)].width = 2

        # ── Write battle blocks: 3 per row ────────────────────────────
        current_row  = 4   # start after title rows + 1 blank
        max_rows_in_band = 0

        for gi, group in enumerate(display_groups):
            pos_in_band = gi % BATTLES_PER_ROW  # 0, 1, 2
            start_col   = 1 + pos_in_band * (BATTLE_COLS + GAP_COLS)

            rows_used = write_battle(gi, group, current_row, start_col)
            max_rows_in_band = max(max_rows_in_band, rows_used)

            # After every 3rd battle (or last battle), advance the row
            if pos_in_band == BATTLES_PER_ROW - 1 or gi == len(display_groups) - 1:
                current_row      += max_rows_in_band + 1  # +1 blank gap between bands
                max_rows_in_band  = 0

        # ── Stream ────────────────────────────────────────────────────
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx_bytes = buf.getvalue()

        safe_name = (sr.name or 'roster').replace(' ', '_')
        response = current_app.response_class(
            xlsx_bytes,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response.headers['Content-Disposition'] = f'attachment; filename="{safe_name}.xlsx"'
        response.headers['Content-Length'] = len(xlsx_bytes)
        return response

    except Exception as e:
        current_app.logger.error(f'Export error: {traceback.format_exc()}')
        flash(f'Export failed: {e}', 'error')
        return redirect(url_for('main.saved_roster_view', roster_id=roster_id))


# ---------------------------------------------------------------------------
# Manual Roster Composer
# ---------------------------------------------------------------------------

MANUAL_TIER_LABELS = ['8', '8', '8', '4', '4', '4', '2', '2', '2', '1', '1', '1']
MANUAL_TIER_NAMES  = {'8': 'manual_tier_8', '4': 'manual_tier_4',
                      '2': 'manual_tier_2', '1': 'manual_tier_1'}


@main_bp.route('/manual-roster', methods=['GET', 'POST'])
@login_required
def manual_roster_select():
    tr = t()
    if current_user.is_admin():
        all_uploads = Upload.query.order_by(Upload.uploaded_at.desc()).all()
    else:
        all_uploads = Upload.query.filter(
            (Upload.user_id == current_user.id) | (Upload.is_shared == True)
        ).order_by(Upload.uploaded_at.desc()).all()

    if request.method == 'POST':
        selected = request.form.getlist('upload_ids')
        if not selected:
            flash(tr['manual_error_no_files'], 'error')
            return redirect(request.url)
        if len(selected) > 3:
            flash(tr['manual_error_too_many'], 'error')
            return redirect(request.url)
        ids_str = ','.join(selected[:3])
        return redirect(url_for('main.manual_roster_compose', ids=ids_str))

    return render_template('manual_roster_select.html', uploads=all_uploads)


@main_bp.route('/manual-roster/compose')
@login_required
def manual_roster_compose():
    tr = t()
    ids_str = request.args.get('ids', '')
    try:
        upload_ids = [int(i) for i in ids_str.split(',') if i.strip()]
    except ValueError:
        flash(tr['manual_error_no_files'], 'error')
        return redirect(url_for('main.manual_roster_select'))

    all_players = []
    columns_used = None
    source_files = []

    for uid in upload_ids:
        upload_rec = _get_accessible_upload(uid)
        if not upload_rec:
            continue
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], upload_rec.filename)
        try:
            df = parse_txt_to_df(filepath)
        except Exception:
            continue

        attendance = Attendance.query.filter_by(upload_id=uid).first()
        confirmed_map = json.loads(attendance.confirmed_rows) if attendance else {}
        power_map  = json.loads(attendance.power_rows  if attendance and attendance.power_rows  else '{}')
        absent_map = json.loads(attendance.absent_rows if attendance and attendance.absent_rows else '{}')

        cols = list(df.columns)
        if columns_used is None:
            columns_used = cols[1:-1]  # strip first and last

        rows = df.fillna('').to_dict(orient='records')
        for i, row in enumerate(rows):
            row['_confirmed'] = confirmed_map.get(str(i), False)
            row['_power']     = power_map.get(str(i), False)
            row['_absent']    = absent_map.get(str(i), False)
            row['_source']    = upload_rec.original_filename
            all_players.append(row)
        source_files.append(upload_rec.original_filename)

    # Deduplicate by name, keep highest reso
    seen = {}
    for p in sorted(all_players, key=_reso, reverse=True):
        name = p.get('Nazwa', '')
        if name and name not in seen:
            seen[name] = p
    pool = list(seen.values())

    # Clean internal keys for template, preserving useful ones
    clean_pool = []
    for p in pool:
        pc = {k: v for k, v in p.items() if not k.startswith('_')}
        pc['confirmed']   = p.get('_confirmed', False)
        pc['power']       = p.get('_power', False)
        pc['absent']      = p.get('_absent', False)
        pc['source_file'] = p.get('_source', '')
        clean_pool.append(pc)

    config = {
        'battle_type':   'Manual',
        'source_files':  source_files,
        'tier_labels':   MANUAL_TIER_LABELS,
        'num_battles':   12,
    }

    source_color_map = {sf: i for i, sf in enumerate(source_files)}
    _ol_map = {'Only confirmed online': tr['opt_online_only'], 'Prioritize Online': tr['opt_prioritize_online'],
               'All players': tr['opt_all_players'], 'Exclude Absent': tr['opt_exclude_absent']}
    online_label = ' + '.join(_ol_map.get(o, o) for o in (online_only if isinstance(online_only, list) else [online_only]))

    # Store heavy data server-side to avoid large hidden form fields
    import pickle
    tmp_path = os.path.join(current_app.config['UPLOAD_FOLDER'],
                            f'_manual_tmp_{current_user.id}.pkl')
    with open(tmp_path, 'wb') as f:
        pickle.dump({
            'columns':     columns_used or [],
            'player_pool': clean_pool,
            'config':      config,
        }, f)

    return render_template('manual_roster_compose.html',
                           pool=clean_pool,
                           columns=columns_used or [],
                           source_files=source_files,
                           source_color_map=source_color_map,
                           ids_str=ids_str,
                           tier_labels=MANUAL_TIER_LABELS,
                           tier_names=MANUAL_TIER_NAMES,
                           config_json=json.dumps(config))


@main_bp.route('/manual-roster/save', methods=['POST'])
@login_required
def manual_roster_save():
    from models import SavedRoster
    import traceback
    tr = t()

    import pickle
    roster_state_json = request.form.get('roster_state', '')
    roster_name       = request.form.get('roster_name', '').strip()

    if not roster_state_json:
        flash('No roster data received — please try again.', 'error')
        return redirect(url_for('main.manual_roster_select'))

    # Load heavy data from temp file
    tmp_path = os.path.join(current_app.config['UPLOAD_FOLDER'],
                            f'_manual_tmp_{current_user.id}.pkl')
    if not os.path.exists(tmp_path):
        flash('Session expired — please reopen the composer.', 'error')
        return redirect(url_for('main.manual_roster_select'))

    try:
        with open(tmp_path, 'rb') as f:
            tmp = pickle.load(f)
        columns    = tmp.get('columns', [])
        pool       = tmp.get('player_pool', [])
        config     = tmp.get('config', {})

        groups = json.loads(roster_state_json)
        today  = date.today().strftime('%Y-%m-%d')
        name   = roster_name if roster_name else f"{current_user.username}_manual_{today}"

        sr = SavedRoster(
            name=name,
            created_by=current_user.id,
            battle_type='Manual',
            config=json.dumps(config),
            groups_data=json.dumps(groups),
            columns_data=json.dumps(columns),
            player_pool=json.dumps(pool),
            overrides='{}',
        )
        db.session.add(sr)
        db.session.commit()
        flash(f'{tr["manual_saved"]} "{name}"', 'success')
        return redirect(url_for('main.saved_roster_view', roster_id=sr.id))
    except Exception as e:
        current_app.logger.error(f'manual_roster_save error: {traceback.format_exc()}')
        flash(f'Save failed: {e}', 'error')
        return redirect(url_for('main.manual_roster_select'))


# ---------------------------------------------------------------------------
# Fast Track Generator
# ---------------------------------------------------------------------------

@main_bp.route('/fast-track', methods=['GET', 'POST'])
@login_required
def fast_track():
    tr = t()
    if request.method == 'POST':
        file = request.files.get('csv_file')
        clan = request.form.get('clan', '').strip()

        if not file or file.filename == '':
            flash(tr['fast_track_error_no_file'], 'error')
            return redirect(request.url)
        if not file.filename.lower().endswith('.txt'):
            flash(tr['upload_error_wrong_type'], 'error')
            return redirect(request.url)
        if clan not in CLAN_OPTIONS:
            flash(tr['upload_error_no_clan'], 'error')
            return redirect(request.url)

        # Save file to uploads
        filename    = f"{uuid.uuid4().hex}.txt"
        filepath    = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        try:
            df = parse_txt_to_df(filepath)
        except Exception as e:
            os.remove(filepath)
            flash(f'{tr["fast_track_error_parse"]}: {e}', 'error')
            return redirect(request.url)

        row_count    = len(df)
        today        = date.today().strftime('%Y-%m-%d')
        display_name = f"{clan}_{today}.txt"

        upload_rec = Upload(
            filename=filename,
            original_filename=display_name,
            user_id=current_user.id,
            is_shared=False,
            row_count=row_count,
            columns=json.dumps(list(df.columns)),
        )
        db.session.add(upload_rec)
        db.session.commit()

        # Build player pool with fixed Fast Track settings
        cols       = list(df.columns)
        col_data   = cols[1:-1]
        rows       = df.fillna('').to_dict(orient='records')
        all_players = []
        for i, row in enumerate(rows):
            row['_confirmed'] = False   # no attendance data yet
            row['_power']     = False
            row['_source']    = display_name
            all_players.append(row)

        # Deduplicate
        seen = {}
        for p in sorted(all_players, key=_reso, reverse=True):
            name = p.get('Nazwa', '')
            if name and name not in seen:
                seen[name] = p
        all_players = list(seen.values())

        # Fixed Fast Track options
        FT_BATTLE_TYPE  = 'Clan Battle'
        FT_CLAN_MODE    = '8 4 2 1'
        FT_PRIORITY     = 'Class'
        FT_ONLINE       = 'Prioritize Online'
        FT_DIST         = 'Even distribution'
        FT_NUM_BATTLES  = 12
        FT_POWER_PLAYER = 'Yes — apply 20% boost'

        # Class filter — all classes
        active_classes = ALL_CLASSES
        all_players    = [p for p in all_players if p.get('Klasa', '') in active_classes]

        # PowerPlayer boost — mark all as power (no attendance, use all)
        # (no boost since no power_rows data — just skip boost silently)

        pool = _apply_online_filter(all_players, FT_ONLINE, FT_NUM_BATTLES * 8)

        groups, tier_labels = _build_groups(
            pool, FT_NUM_BATTLES, FT_PRIORITY, FT_DIST, clan_mode=FT_CLAN_MODE
        )

        # Preserve source_file
        for g in groups:
            for p in g:
                if p.get('_source'):
                    p['source_file'] = p['_source']
        for p in all_players:
            if p.get('_source'):
                p['source_file'] = p['_source']

        def _clean(p):
            return {k: v for k, v in p.items() if not k.startswith('_')}

        clean_groups   = [[_clean(p) for p in g] for g in groups]
        full_pool_clean = [_clean(p) for p in all_players]

        config = {
            'battle_type':    FT_BATTLE_TYPE,
            'clan_mode':      FT_CLAN_MODE,
            'priority':       FT_PRIORITY,
            'online_only':    FT_ONLINE,
            'distribution':   FT_DIST,
            'num_battles':    FT_NUM_BATTLES,
            'active_classes': active_classes,
            'source_files':   [display_name],
            'tier_labels':    tier_labels,
            'power_player':   FT_POWER_PLAYER,
        }

        from models import SavedRoster
        name = f"{current_user.username}_{today}"
        sr = SavedRoster(
            name=name,
            created_by=current_user.id,
            battle_type=FT_BATTLE_TYPE,
            config=json.dumps(config),
            groups_data=json.dumps(clean_groups),
            columns_data=json.dumps(col_data),
            player_pool=json.dumps(full_pool_clean),
            overrides='{}',
        )
        db.session.add(sr)
        db.session.commit()

        # Build and stream Excel inline (same logic as saved_roster_export)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from io import BytesIO
            import traceback as _tb

            black='000000'; white='FFFFFF'; light_gray='F2F2F2'
            mid_gray='CCCCCC'; dark_gray='444444'; hdr_fill='2A2A2A'; accent_dk='3A6B00'
            tier_colours={'8':('FFF0E0','7A3A00'),'4':('E8F5E8','1A5C1A'),
                          '2':('E0EEF8','0A3A6A'),'1':('FCE8E8','7A1A1A')}
            class_colours={
                'Nekromanta':('EEF0FF','3A3A8A'),'Barbarzyca':('FFF0E0','7A3A00'),
                'Barbarzyńca':('FFF0E0','7A3A00'),'Łowca demonów':('EAFAEA','1A5C1A'),
                'Rycerz krwi':('FFE8E8','7A1A1A'),'Czarownik':('F5E8FF','6A1A8A'),
                'Mnich':('E8F8FF','0A5A7A'),'Krzyżowiec':('FFFAE0','6A5A00'),
                'Sztormiciel':('E8F0FF','1A3A8A'),'Druid':('E8FFEE','0A5A2A'),
            }
            def sol(h): return PatternFill('solid', start_color=h, fgColor=h)
            def bdr(c=mid_gray):
                s=Side(style='thin',color=c); return Border(left=s,right=s,top=s,bottom=s)

            BATTLES_PER_ROW=3; GAP_COLS=1; BATTLE_COLS=1+len(col_data)
            total_cols=BATTLES_PER_ROW*BATTLE_COLS+(BATTLES_PER_ROW-1)*GAP_COLS

            wb=Workbook(); ws=wb.active; ws.title=name[:31]
            ws.sheet_view.showGridLines=False

            # Title
            ws.row_dimensions[1].height=22
            c=ws.cell(1,1,name)
            c.font=Font(name='Arial',bold=True,size=13,color=accent_dk)
            c.fill=sol(light_gray); c.alignment=Alignment(horizontal='left',vertical='center')
            ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=total_cols)
            ws.row_dimensions[2].height=13
            c=ws.cell(2,1,f'{FT_BATTLE_TYPE} · {FT_CLAN_MODE} · {FT_PRIORITY} · {FT_DIST}')
            c.font=Font(name='Arial',size=8,color=dark_gray,italic=True)
            c.fill=sol(light_gray); c.alignment=Alignment(horizontal='left',vertical='center')
            ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=total_cols)

            # Column widths
            col_widths=[4]+[max(13,len(c2)+2) for c2 in col_data]
            for block in range(BATTLES_PER_ROW):
                base=1+block*(BATTLE_COLS+GAP_COLS)
                for ci,w in enumerate(col_widths):
                    ws.column_dimensions[get_column_letter(base+ci)].width=w
                if block<BATTLES_PER_ROW-1:
                    ws.column_dimensions[get_column_letter(base+BATTLE_COLS)].width=2

            def write_battle(gi, group, start_row, start_col):
                tier=tier_labels[gi] if tier_labels and gi<len(tier_labels) else None
                if tier and tier_labels:
                    same_tier_before = sum(1 for t2 in tier_labels[:gi] if t2 == tier)
                    pos_in_tier = same_tier_before + 1
                    bl = f'Battle {gi+1} ({pos_in_tier} for {tier}pts)'
                else:
                    bl = f'Battle {gi+1}'
                tf,tt=tier_colours.get(tier,(light_gray,dark_gray)) if tier else (light_gray,dark_gray)
                end_col=start_col+BATTLE_COLS-1
                ws.row_dimensions[start_row].height=18
                c=ws.cell(start_row,start_col,bl)
                c.font=Font(name='Arial',bold=True,size=9,color=tt); c.fill=sol(tf)
                c.alignment=Alignment(horizontal='center',vertical='center'); c.border=bdr()
                ws.merge_cells(start_row=start_row,start_column=start_col,end_row=start_row,end_column=end_col)
                sub=start_row+1; ws.row_dimensions[sub].height=16
                for ci2,hdr in enumerate(['#']+col_data):
                    c=ws.cell(sub,start_col+ci2,hdr)
                    c.font=Font(name='Arial',bold=True,size=8,color=white); c.fill=sol(hdr_fill)
                    c.alignment=Alignment(horizontal='center' if ci2==0 else 'left',vertical='center')
                    c.border=bdr(hdr_fill)
                for pi,player in enumerate(group):
                    dr=start_row+2+pi; ws.row_dimensions[dr].height=15
                    rbg=white if pi%2==0 else light_gray
                    c=ws.cell(dr,start_col,pi+1)
                    c.font=Font(name='Arial',size=8,color=dark_gray); c.fill=sol(rbg)
                    c.alignment=Alignment(horizontal='center',vertical='center'); c.border=bdr()
                    for ci2,col2 in enumerate(col_data,1):
                        val=str(player.get(col2,''or''))
                        c=ws.cell(dr,start_col+ci2,val)
                        c.fill=sol(rbg); c.alignment=Alignment(horizontal='left',vertical='center')
                        c.font=Font(name='Arial',size=8,color=black); c.border=bdr()
                        if col2=='Nazwa': c.font=Font(name='Arial',size=8,bold=True,color=black)
                        elif col2=='Klasa':
                            bg2,fg2=class_colours.get(val,(rbg,dark_gray))
                            c.fill=sol(bg2); c.font=Font(name='Arial',size=8,bold=True,color=fg2)
                            c.alignment=Alignment(horizontal='center',vertical='center')
                        elif col2 in ('Rezonowanie','Ranking udziału','Poziom'):
                            c.alignment=Alignment(horizontal='center',vertical='center')
                try:
                    avg=round(sum(int(p.get('Rezonowanie',0)or 0) for p in group)/len(group)) if group else 0
                except: avg=0
                fr=start_row+2+len(group); ws.row_dimensions[fr].height=12
                c=ws.cell(fr,start_col,f'{len(group)} players · avg reso {avg}')
                c.font=Font(name='Arial',size=7,color=dark_gray,italic=True); c.fill=sol(light_gray)
                c.alignment=Alignment(horizontal='left',vertical='center'); c.border=bdr()
                ws.merge_cells(start_row=fr,start_column=start_col,end_row=fr,end_column=end_col)
                return 2+len(group)+1

            cur_row=4; max_band=0
            for gi,group in enumerate(clean_groups):
                pos=gi%BATTLES_PER_ROW
                sc=1+pos*(BATTLE_COLS+GAP_COLS)
                used=write_battle(gi,group,cur_row,sc)
                max_band=max(max_band,used)
                if pos==BATTLES_PER_ROW-1 or gi==len(clean_groups)-1:
                    cur_row+=max_band+1; max_band=0

            buf=BytesIO(); wb.save(buf); buf.seek(0)
            xlsx_bytes=buf.getvalue()
            safe_name=name.replace(' ','_')
            # Store excel bytes in session-linked temp file, redirect to download endpoint
            import pickle as _pkl
            ft_tmp = os.path.join(current_app.config['UPLOAD_FOLDER'], f'_ft_xlsx_{current_user.id}.pkl')
            with open(ft_tmp, 'wb') as _f:
                _pkl.dump({'bytes': xlsx_bytes, 'name': safe_name}, _f)
            session['ft_roster_id'] = sr.id
            flash(tr['fast_track_success'], 'success')
            return redirect(url_for('main.fast_track_download'))

        except Exception as e:
            current_app.logger.error(f'fast_track export error: {_tb.format_exc()}')
            flash(f'{tr["fast_track_error_parse"]}: {e}', 'error')
            return redirect(url_for('main.saved_roster_view', roster_id=sr.id))

    return render_template('fast_track.html', clan_options=CLAN_OPTIONS)


@main_bp.route('/fast-track/download')
@login_required
def fast_track_download():
    """Intermediate page: triggers /fast-track/excel in iframe, then redirects."""
    roster_id   = session.get('ft_roster_id')
    redirect_url = url_for('main.saved_roster_view', roster_id=roster_id) if roster_id                    else url_for('main.saved_rosters_list')
    excel_url   = url_for('main.fast_track_excel')

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Fast Track</title>
<style>body{{font-family:sans-serif;padding:60px;color:#888;background:#0f0f0f;}}</style>
</head><body>
<p>⚡ Downloading Excel and redirecting…</p>
<iframe src="{excel_url}" style="display:none"></iframe>
<script>setTimeout(function(){{ window.location="{redirect_url}"; }}, 1500);</script>
</body></html>"""
    return html


@main_bp.route('/fast-track/excel')
@login_required
def fast_track_excel():
    """Serves the Excel file directly."""
    import pickle as _pkl
    ft_tmp = os.path.join(current_app.config['UPLOAD_FOLDER'], f'_ft_xlsx_{current_user.id}.pkl')

    if not os.path.exists(ft_tmp):
        return '', 204

    with open(ft_tmp, 'rb') as _f:
        data = _pkl.load(_f)
    os.remove(ft_tmp)

    session.pop('ft_roster_id', None)   # clean up after served

    xlsx_bytes = data['bytes']
    safe_name  = data['name']

    resp = current_app.response_class(
        xlsx_bytes,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp.headers['Content-Disposition'] = f'attachment; filename="{safe_name}.xlsx"'
    resp.headers['Content-Length'] = len(xlsx_bytes)
    return resp
